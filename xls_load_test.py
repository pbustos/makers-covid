from openpyxl import load_workbook
from unidecode import unidecode

wb = load_workbook(filename='vextremadura.xlsx',
                   read_only=True)

ws = wb['Caceres']

# Read the cell values into a list of lists
data_rows = []

def remove_special_chars(text):
    return unidecode.unidecode(text)

headers = ws['B4':'H4']
header_values = list(map(lambda x: clean_and_map_header(x.value, headers_mapping), headers))
json = {}
for row in ws['B5':'H1000']:
    row_values = list(map(lambda x: remove_special_chars(x.value.strip()), row))
    for row in chunks(data, len(char_range(column_range))):
        row_values = list(map(lambda x: remove_special_chars(x.value.strip()), row))
        if row_values[0] != '':
            row_values = reasign_type(types_map, row_values)
            row_dict = dict(zip(header_values, row_values))
            row_dict["Localidad"] = self.sheet.title
            json[row_values[0]] = (row_dict)
    return json

print(data_rows)