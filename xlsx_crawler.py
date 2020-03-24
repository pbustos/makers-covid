import re
import time

import yaml
from openpyxl import load_workbook
from unidecode import unidecode
import utils
import requests

with open(r'config.yml', encoding='utf8') as file:
    config = yaml.load(file, Loader=yaml.FullLoader)
    SPREAD_SHEET_URL = config["SPREAD_SHEET_URL"]
    SHEET_DATA = config["SHEET_DATA"]
    SCOPES = config["SCOPES"]
    SOCKETIO_PORT = config["SOCKETIO_PORT"]


class XLSXCrawler:
    def __init__(self):
        self.__spreadsheet = None
        self.__last_time = time.time()

    @property
    def spreadsheet(self):
        elapsed_time = time.time() - self.__last_time
        if self.__spreadsheet is None or elapsed_time > 1:
            self.__init_spreadsheet()
        return self.__spreadsheet

    def __init_spreadsheet(self):
        id_search = re.search("([^\/]{20,})", config["SPREAD_SHEET_URL"])
        assert id_search, "NO ID FOUND IN URL"
        id = id_search.group(0)
        response = requests.get(
            'https://docs.google.com/spreadsheet/ccc?key='+id+'&output=xlsx')
        assert response.status_code == 200, 'Wrong status code'
        with open('vextremadura.xlsx', 'wb') as outfile:
            outfile.write(response.content)
        self.__spreadsheet = load_workbook(filename='vextremadura.xlsx', read_only=True)
        self.__last_time = time.time()

    def change_sheet(self, sheet):
        self.worksheet = self.spreadsheet['Caceres']

    def update_stock(self):
        stock_column_range = SHEET_DATA[self.worksheet.title]["STOCK_COLUMN_RANGE"]
        stock_initial_column = SHEET_DATA[self.worksheet.title]["STOCK_INITIAL_ROW"]
        stock_headers_map = SHEET_DATA[self.worksheet.title]["STOCK_HEADERS_MAP"]
        stock_column_types = SHEET_DATA[self.worksheet.title]["STOCK_COLUMN_TYPES"]
        return self.update_subtable(stock_column_range,stock_initial_column, stock_headers_map, stock_column_types)

    def update_demand(self):
        demand_column_range = SHEET_DATA[self.worksheet.title]["DEMAND_COLUMN_RANGE"]
        demand_initial_column = SHEET_DATA[self.worksheet.title]["DEMAND_INITIAL_ROW"]
        demand_headers_map = SHEET_DATA[self.worksheet.title]["DEMAND_HEADERS_MAP"]
        demand_column_types = SHEET_DATA[self.worksheet.title]["DEMAND_COLUMN_TYPES"]
        return self.update_subtable(demand_column_range, demand_initial_column, demand_headers_map, demand_column_types)

    def update_subtable(self, column_range, initial_row, headers_mapping, types_map):
        headers = self.worksheet[column_range[0]+str(initial_row):column_range[-1]+str(initial_row)]
        header_values = list(map(lambda x: utils.clean_and_map_header(x.value, headers_mapping), headers[0]))
        data = self.worksheet[column_range[0]+str(initial_row+1):column_range[-1]+str(1000)]
        json = {}
        for row in data:
            row_values = list(map(lambda x: utils.remove_special_chars(x.value), row))
            if row_values[0] != '':
                row_values = utils.reasign_type(types_map, row_values)
                row_dict = dict(zip(header_values,row_values))
                row_dict["Localidad"] = self.worksheet.title
                json[row_values[0]]= (row_dict)
        return json

    def get_worksheet_data(self, worksheet):
        self.change_sheet(worksheet)
        final_json = {}
        final_json["demand"] = self.update_demand()
        final_json["makers"] = self.update_stock()
        return final_json

# wb = load_workbook(filename='vextremadura.xlsx',
#                    read_only=True)
#
# ws = wb['Caceres']
#
# # Read the cell values into a list of lists
# data_rows = []
#
# def remove_special_chars(text):
#     return unidecode.unidecode(text)
#
# headers = ws['B4':'H4']
# header_values = list(map(lambda x: utils.clean_and_map_header(x.value, headers_mapping), headers))
# json = {}
# for row in ws['B5':'H1000']:
#
#     row_values = list(map(lambda x: remove_special_chars(x.value.strip()), row))
#     if row_values[0] != '':
#         row_values = utils.reasign_type(types_map, row_values)
#         row_dict = dict(zip(header_values, row_values))
#         row_dict["Localidad"] = self.sheet.title
#         json[row_values[0]] = (row_dict)
#
#
# print(data_rows)

if __name__ == '__main__':
    a = XLSXCrawler()
    print(a.get_worksheet_data('Caceres'))